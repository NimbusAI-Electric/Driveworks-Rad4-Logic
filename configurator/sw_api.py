"""
sw_api.py
=========
SolidWorks COM API bridge for the RAD4 configurator.

Uses template copying and reference redirection (ReplaceReferencedDocument)
to generate configured mirror models, assemblies, drawings, and BOMs without
altering the original template files. Supports custom/decimal sizes by dynamically
mapping to the closest standard templates, copying the structural extrusions (72239)
and diffusers (64792), updating their length dimensions, and referencing them.
"""

import os
import shutil
import glob
import logging
import re
from pathlib import Path
import win32com.client as win32
import pythoncom
from rad4_engine import RAD4Inputs, RAD4Result, DWConstantVault

log = logging.getLogger(__name__)

# --- Note Text Constants from Claude rules ---
SPEC_BASE = (
    "BRING MC CABLE TO ENCLOSURE. INSERT GROUND WIRE IN GROUNDED CONNECTOR. "
    "INSERT HOT AND NEUTRAL WIRE INTO LUMINAIRE DISCONNECT. "
    "NO ELECTRICAL BOX REQUIRED. ELECTRICAL POWER SHOULD BE CONTROLLED "
    "BY A WALL SWITCH (BY OTHERS). MIRROR SHOULD BE MOUNTED TO A "
    "MECHANICALLY SOUND SURFACE SUCH AS WALL STUDS TO SUPPORT ITS WEIGHT."
)
SPEC_DIM_0_10V = (
    "BRING MC CABLE TO ENCLOSURE. INSERT GROUND WIRE IN GROUNDED CONNECTOR. "
    "INSERT HOT AND NEUTRAL WIRE INTO LUMINAIRE DISCONNECT. "
    "0-10V CONTROL WIRES ARE BROUGHT IN THROUGH THE SECOND KNOCKOUT. "
    "NO ELECTRICAL BOX REQUIRED. ELECTRICAL POWER SHOULD BE CONTROLLED "
    "BY A WALL SWITCH (BY OTHERS). MIRROR SHOULD BE MOUNTED TO A "
    "MECHANICALLY SOUND SURFACE SUCH AS WALL STUDS TO SUPPORT ITS WEIGHT."
)
SPEC_CORD = (
    "PLUG FIXTURE INTO RECEPTACLE LOCATED IN WALL BEHIND MIRROR. "
    "RECEPTACLES SHOULD BE CONTROLLED BY WALL SWITCHES (BY OTHERS). "
    "MIRROR SHOULD BE MOUNTED TO A MECHANICALLY SOUND SURFACE SUCH AS "
    "WALL STUDS TO SUPPORT ITS WEIGHT."
)
SPEC_HANGING = (
    "BRING MC CABLE TO DRIVER ENCLOSURE EITHER DIRECTLY FROM BEHIND INTO "
    "KNOCKOUT OR PROVIDE (30\" MAX) WHIP TO SIDE KNOCKOUT. INSERT GROUND "
    "WIRE IN GROUNDED CONNECTOR. INSERT HOT AND NEUTRAL WIRE INTO "
    "LUMINAIRE DISCONNECT. LOW VOLTAGE CONTROL WIRES ARE BROUGHT IN "
    "THROUGH THE SECOND KNOCKOUT. NO ELECTRICAL BOX REQUIRED. ELECTRICAL "
    "POWER SHOULD BE CONTROLLED BY A WALL SWITCH (BY OTHERS). HANGING "
    "BRACKET SHOULD BE MOUNTED TO A MECHANICALLY SOUND SURFACE SUCH AS "
    "WALL STUDS TO SUPPORT FIXTURE WEIGHT."
)
SPEC_WG3_TAIL = " MAIN LIGHTS AND WALL GLOW OPERATE TOGETHER."

ATTN_NEC = (
    "THIS PRODUCT MUST BE CONNECTED TO EARTH GROUND IN ACCORDANCE "
    "WITH NEC CODE 250.20 (B). IMPROPER GROUND CAN RESULT IN "
    "IRREGULAR FUNCTION OF THE UNIT."
)
ATTN_GFCI = (
    "ELECTRIC MIRROR RECOMMENDS A HARDWIRED INSTALLATION AS THE PREFERRED "
    "METHOD OF INSTALLATION FOR ALL LIGHTED MIRROR PRODUCTS. PRIOR TO THE "
    "DELIVERY OF THIS CORD CONNECTED LUMINAIRE, WE RECOMMEND THAT YOU CONTACT "
    "YOUR LOCAL ELECTRICAL INSPECTOR TO REVIEW THE PLANNED CONDITIONS FOR THE "
    "INSTALLATION OF THIS PRODUCT. THIS WILL ENSURE COMPLIANCE WITH THE LOCAL "
    "ELECTRICAL CODE. IF A GFCI CIRCUIT IS REQUIRED, INSTALL A NON-GFCI OUTLET "
    "BEHIND THE MIRROR. THIS OUTLET MUST BE WIRED FROM THE LOAD SIDE OF AN "
    "ACCESSIBLE GFCI OUTLET. THIS WILL PREVENT HAVING TO REMOVE THE MIRROR TO "
    "RESET THE GFCI RECEPTACLE. THE LOAD ON THIS GFCI CIRCUIT SHOULD BE "
    "CAREFULLY CONSIDERED TO PREVENT UNINTENDED TRIPPING OF THE GFCI. MUST BE "
    "INSTALLED IN ACCORDANCE WITH ALL NATIONAL AND LOCAL ELECTRICAL CODES. "
    "ELECTRIC MIRROR IS NOT RESPONSIBLE FOR COMPATIBILITY OF A GFCI CIRCUIT "
    "WITH OUR PRODUCTS. CONTACT THE GFCI MANUFACTURER TO ENSURE COMPATIBILITY."
)

DIM_TRIAC = (
    "TO ENSURE PROPER OPERATION OF THIS DIMMABLE PRODUCT IT IS IMPORTANT TO "
    "SELECT A COMPATIBLE DIMMING SWITCH. THIS LUMINAIRE REQUIRES A COMPATIBLE "
    "FORWARD PHASE LINE DIMMER SWITCH. CONTACT THE CONTROLLER MANUFACTURER "
    "TO CONFIRM COMPATIBILITY WITH THIS PRODUCT. MUST BE INSTALLED IN "
    "ACCORDANCE WITH ALL NATIONAL AND LOCAL ELECTRICAL CODES. ELECTRIC "
    "MIRROR IS NOT RESPONSIBLE FOR DIMMER SWITCH COMPATIBILITY. THIS PRODUCT "
    "USES: SMT-024-096VTSP TRIAC PHASE DIMMING DRIVER."
)
DIM_0_10V = (
    "TO ENSURE PROPER OPERATION OF THIS DIMMABLE PRODUCT, IT IS IMPORTANT TO "
    "SELECT A COMPATIBLE DIMMING SWITCH. THIS LUMINAIRE REQUIRES A 0-10V "
    "ELECTRONIC DIMMER SWITCH. ELECTRIC MIRROR IS NOT RESPONSIBLE FOR DIMMER "
    "SWITCH COMPATIBILITY. MUST BE INSTALLED IN ACCORDANCE WITH ALL NATIONAL "
    "AND LOCAL ELECTRICAL CODES."
)

DEF_KEEN_DISC = (
    "KEEN UNIT CONTROLS THE LIGHTING ONLY SEPERATE WALL SWITCH "
    "IS REQUIRED TO CONTROL DEFOGGER POWER"
)

def parse_cpn_options(cpn: str) -> set[str]:
    parts = [p.strip().upper() for p in cpn.split("-")]
    # Find dimension token index
    dim_idx = -1
    for i, p in enumerate(parts):
        if "X" in p and re.search(r"\d+\.\d+X\d+\.\d+", p):
            dim_idx = i
            break
    if dim_idx == -1:
        return set()
    
    finish_idx = dim_idx + 1
    # Options are all parts after finish_idx, except the last one (which is CCT)
    if finish_idx < len(parts) - 1:
        return set(parts[finish_idx + 1 : -1])
    return set()

def _get_drawing_notes_from_opts(w: float, h: float, opts: set[str]) -> dict:
    area = w * h
    has_d1 = "D1" in opts
    has_d2 = "D2" in opts
    has_df = "DF" in opts
    has_dfx = "DFX" in opts
    has_any_keen = bool(opts & {"KG", "KG2", "KD", "KC"})
    has_any_clock = bool(opts & {"CK2", "CK3"})
    has_cc2 = "CC2" in opts
    has_wg3 = "WG3" in opts
    has_no = "NO" in opts
    has_277V = "277V" in opts

    notes = {}

    # SPECIFICATION
    if has_cc2:
        spec = SPEC_CORD
    elif has_any_clock:
        spec = SPEC_HANGING
    elif has_d1 or has_df or has_dfx:
        spec = SPEC_DIM_0_10V
    else:
        spec = SPEC_BASE

    if has_wg3 or has_no:
        spec += SPEC_WG3_TAIL
    notes["SPECIFICATION"] = spec

    # ATTENTION
    if has_cc2:
        notes["ATTENTION"] = ATTN_GFCI
    elif has_any_keen:
        notes["ATTENTION"] = ATTN_NEC

    # DEFOGGER
    if has_df or has_dfx:
        if has_277V:
            w_str = "20W @ 24V"
        elif has_dfx:
            w_str = "25W @ 120V"
        elif area < 900:
            w_str = "15W @ 120V"
        elif area < 1900:
            w_str = "25W @ 120V"
        elif area < 2465:
            w_str = "50W @ 120V"
        else:
            w_str = "100W @ 120V"
        notes["DEFOGGER"] = f"VOLTAGE / WATTAGE: {w_str}"

    # DEFOGGER DISCLAIMER (KEEN)
    if has_any_keen and (has_df or has_dfx):
        notes["DEFOGGER DISCLAIMER (KEEN)"] = DEF_KEEN_DISC

    # DIMMING
    if has_d2:
        notes["DIMMING"] = DIM_TRIAC
    elif has_d1:
        notes["DIMMING"] = DIM_0_10V

    return notes


# COM Constants
swDocPART = 1
swDocASSEMBLY = 2
swDocDRAWING = 3
swSaveAsCurrentVersion = 0
swSaveAsOptions_Silent = 1

def _copy_file_writable(src: Path, dest: Path):
    try:
        src_abs = Path(src).resolve()
        dest_abs = Path(dest).resolve()
        if src_abs == dest_abs:
            if dest_abs.exists():
                try:
                    os.chmod(str(dest_abs), 0o666)
                except Exception as e:
                    log.warning(f"Could not make destination file writable {dest_abs}: {e}")
            return
    except Exception as e:
        log.warning(f"Error checking file path identity for {src} and {dest}: {e}")

    if dest.exists():
        try:
            os.chmod(str(dest), 0o666)
            dest.unlink()
        except Exception as e:
            log.warning(f"Could not prepare destination file {dest} for copying: {e}")
    
    shutil.copyfile(str(src), str(dest))
    try:
        os.chmod(str(dest), 0o666)
    except Exception as e:
        log.warning(f"Could not make destination file writable {dest}: {e}")

def find_closest_template(directory: Path, pattern: str, width: float, height: float, exclude_name: str = None) -> Path:
    """
    Finds the template file in the directory that is closest in size to (width, height).
    Prioritizes option-heavy templates by sorting by filename length descending.
    """
    files = list(directory.glob(pattern))
    templates = []
    
    # Exclude generated/released files to avoid using configured mirrors as templates
    for f in files:
        name = f.name
        name_upper = name.upper()
        if exclude_name and exclude_name.upper() in name_upper:
            continue
        # If it is a drawing or Excel BOM template, bypass the checks since drawing/BOM templates typically contain finish and color codes
        if not (name_upper.endswith(".SLDDRW") or name_upper.endswith("BOM.XLSX")):
            # Release files typically contain mount types or lighting codes
            if any(token in name_upper for token in ["-RM-", "-SM-", "-LO-", "-SO-", "-LSE-", "-LHE-"]):
                continue
            # Configured files have finish codes (e.g. NK04, CH04, BK05) or color temp (e.g. 30K)
            if re.search(r'-(BK|NK|CH|BN|BR|BZ|GU|WH)\d{2}', name_upper):
                continue
            if re.search(r'-\d{2}K', name_upper):
                continue
        templates.append(f)

    if not templates:
        raise FileNotFoundError(f"No templates matching pattern {pattern} found in {directory}")

    candidates = []
    size_re = re.compile(r'RAD4-(\d+\.\d+)X(\d+\.\d+)', re.IGNORECASE)

    for f in templates:
        m = size_re.search(f.name)
        if m:
            w_val = float(m.group(1))
            h_val = float(m.group(2))
            dist = abs(w_val - width) + abs(h_val - height)
            candidates.append((dist, len(f.name), f))

    if not candidates:
        log.warning(f"Could not parse sizes from templates in {directory}. Falling back to first match.")
        return templates[0]

    # Sort by distance first (ascending), then by length of file name (descending) to prioritize option-heavy templates
    candidates.sort(key=lambda x: (x[0], -x[1]))
    best_file = candidates[0][2]
        
    return best_file


class SolidWorksAPI:
    def __init__(self):
        self.swApp = None
        self._connect()

    def _connect(self):
        try:
            try:
                self.swApp = win32.GetActiveObject("SldWorks.Application")
                log.info("Connected to existing SolidWorks session.")
            except Exception:
                self.swApp = win32.Dispatch("SldWorks.Application.29") # SW2021
                self.swApp.Visible = True
                log.info("Launched new SolidWorks 2021 session.")
        except Exception as e:
            raise RuntimeError(f"Failed to connect to SolidWorks: {e}")

    def generate(self, inp: RAD4Inputs, result: RAD4Result, output_dir: str) -> dict:
        width = inp.UnitWidth
        height = inp.UnitHeight
        cpn = result.CPN
        vault = DWConstantVault
        powerBoxY = 0.0
        hangerY = 0.0

        log.info(f"Starting programmatic generative configuration for: {cpn}")

        # Close all active documents to release file locks before copying
        try:
            self.swApp.CloseAllDocuments(True)
            log.info("Closed all open documents in SolidWorks.")
        except Exception as e:
            log.warning(f"Could not close open documents in SolidWorks: {e}")

        # Define output and Vault directories
        vault_root = Path(vault)
        vault_mirror_dir = vault_root / "Products" / "JS3" / "Mirrors" / "RAD4"
        vault_chassis_dir = vault_root / "Products" / "JS3" / "Assemblies" / "RAD4"
        vault_sales_aid_dir = vault_root / "Products" / "JS3" / "Sales Aid" / "RAD4"

        # Ensure Vault directories exist programmatically
        vault_mirror_dir.mkdir(parents=True, exist_ok=True)
        vault_chassis_dir.mkdir(parents=True, exist_ok=True)
        vault_sales_aid_dir.mkdir(parents=True, exist_ok=True)

        # Ensure output directory (N drive) exists
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        out_mirror = vault_mirror_dir / f"{cpn}-M.SLDASM"
        out_skeleton = vault_mirror_dir / f"{cpn}-SKELETON.SLDPRT"
        out_chassis = vault_chassis_dir / f"{cpn}-C.SLDASM"
        out_drawing = vault_sales_aid_dir / f"{cpn}-SALES-AID.SLDDRW"
        
        # PDFs save to N drive (output_dir)
        pdf_path = os.path.join(output_dir, f"{cpn}-SALES-AID.PDF")
        bom_path = os.path.join(str(vault_mirror_dir), f"{cpn}-BOM.xlsx")
        bom_pdf_path = os.path.join(output_dir, f"{cpn}-BOM.pdf")

        # Local templates inside the git repository configurator/templates/
        templates_dir = Path(__file__).parent / "templates"
        template_skeleton = templates_dir / "RAD4-SKELETON.SLDPRT"
        template_asm = templates_dir / "RAD4-MASTER-ASSEMBLY.SLDASM"

        if not template_skeleton.exists() or not template_asm.exists():
            raise FileNotFoundError(f"Local configurator templates not found in {templates_dir}")

        log.info(f"Copying master skeleton and assembly templates...")
        _copy_file_writable(template_skeleton, out_skeleton)
        _copy_file_writable(template_asm, out_mirror)

        log.info("Updating skeleton reference in the mirror assembly...")
        success_ref = self.swApp.ReplaceReferencedDocument(str(out_mirror), str(template_skeleton), str(out_skeleton))
        log.info(f"ReplaceReferencedDocument (Mirror -> Skeleton) status: {success_ref}")

        # Configure custom frame extrusions and diffuser parts locally in output_dir
        log.info("Configuring custom frame extrusions locally...")
        out_asm_w, out_asm_h, out_asm_hn = self._configure_local_extrusions(vault, output_dir, width, height)

        # Configure custom LED strip locally in output_dir
        log.info("Configuring custom LED strip assembly locally...")
        out_led_asm = self._configure_local_led(vault, output_dir, width, height, result.LEDPN)

        # Load generic chassis template from repository configurator/templates/
        template_chassis = templates_dir / "RAD4-GENERIC-CHASSIS.SLDASM"
        
        if not template_chassis.exists():
            raise FileNotFoundError(f"Generic chassis template not found in {templates_dir}")
        
        # Load generic Sales Aid template drawing from repository configurator/templates/
        template_drawing = templates_dir / "RAD4-GENERIC-SALES-AID.SLDDRW"
        
        if not template_drawing.exists():
            raise FileNotFoundError(f"Generic drawing template not found in {templates_dir}")

        log.info(f"Copying chassis template from vault and local generic drawing template...")
        _copy_file_writable(template_chassis, out_chassis)
        _copy_file_writable(template_drawing, out_drawing)

        # Replace mirror in chassis copy
        ref_mirror_temp = None
        deps = self.swApp.GetDocumentDependencies2(str(out_chassis), True, True, False)
        if deps:
            for i in range(0, len(deps), 2):
                dep_path = deps[i+1] if i+1 < len(deps) else ""
                if dep_path:
                    filename = os.path.basename(dep_path).lower()
                    if filename.startswith("rad4-") and filename.endswith("-m.sldasm"):
                        ref_mirror_temp = dep_path
                        break
        if not ref_mirror_temp:
            m_chassis = re.search(r'RAD4-(\d+\.\d+)X(\d+\.\d+)', template_chassis.name)
            temp_w = float(m_chassis.group(1)) if m_chassis else width
            temp_h = float(m_chassis.group(2)) if m_chassis else height
            mirror_dir = Path(vault) / "Products" / "JS3" / "Mirrors" / "RAD4"
            ref_mirror_temp = str(mirror_dir / f"RAD4-{temp_w:.2f}X{temp_h:.2f}-M.SLDASM")

        success = self.swApp.ReplaceReferencedDocument(str(out_chassis), str(ref_mirror_temp), str(out_mirror))
        log.info(f"ReplaceReferencedDocument (Chassis -> Mirror) replacing {ref_mirror_temp} with {out_mirror}: {success}")

        # Programmatically find and replace assembly dependency in drawing template copy
        drawing_deps = self.swApp.GetDocumentDependencies2(str(out_drawing), True, True, False)
        ref_chassis_temp = None
        if drawing_deps:
            for i in range(0, len(drawing_deps), 2):
                dep_path = drawing_deps[i+1] if i+1 < len(drawing_deps) else ""
                if dep_path and dep_path.lower().endswith(".sldasm"):
                    ref_chassis_temp = dep_path
                    break
        
        if ref_chassis_temp:
            success_draw = self.swApp.ReplaceReferencedDocument(str(out_drawing), str(ref_chassis_temp), str(out_chassis))
            log.info(f"ReplaceReferencedDocument (Drawing -> Chassis) replacing {ref_chassis_temp} with {out_chassis}: {success_draw}")
        else:
            log.warning("Could not automatically resolve assembly dependency in generic drawing template.")

        errors = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        warnings = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

        # A. Open and configure Mirror Assembly (Decoupled generative frame replacement)
        log.info(f"Opening Mirror copy: {out_mirror}")
        swMirror = self.swApp.OpenDoc6(str(out_mirror), swDocASSEMBLY, 1, "", errors, warnings)
        if swMirror:
            self.swApp.ActivateDoc3(swMirror.GetTitle, True, 2, errors)
            
            # 1. Drive skeleton dimensions
            skeleton_comp = None
            components = swMirror.GetComponents(True)
            for comp in components:
                if "skeleton" in comp.Name2.lower():
                    skeleton_comp = comp
                    break

            if skeleton_comp:
                skeleton_model = skeleton_comp.GetModelDoc2
                if skeleton_model:
                    dim_w = skeleton_model.Parameter("Width@SkeletonSketch")
                    dim_h = skeleton_model.Parameter("Height@SkeletonSketch")
                    if dim_w:
                        dim_w.SystemValue = width * 0.0254
                    if dim_h:
                        dim_h.SystemValue = height * 0.0254
                    skeleton_model.ForceRebuild3(False)
                    log.info(f"Driven skeleton sketch size to {width} x {height} inches")
                else:
                    log.warning("Could not get skeleton model document")
            else:
                log.warning("Skeleton component not found in assembly")

            # 2. Drive assembly offset planes and glass sketch
            mirror_dims = {
                "D1@LEFT": (width / 2.0) * 0.0254,
                "D1@RIGHT": (width / 2.0) * 0.0254,
                "D1@TOP": (height / 2.0) * 0.0254,
                "D1@BOTTOM": (height / 2.0) * 0.0254,
                "Glass Width@GlassCutSketch": (width - 1.5354) * 0.0254,
                "Glass Height@GlassCutSketch": (height - 1.5354) * 0.0254,
            }
            for name, val in mirror_dims.items():
                dim = swMirror.Parameter(name)
                if dim:
                    dim.SystemValue = val
                else:
                    log.warning(f"Mirror parameter {name} not found.")

            # 3. Decouple and replace frame and LED components using selection-based ReplaceComponents2
            swAssy = swMirror
            for comp in components:
                name = comp.Name2
                path = comp.GetPathName.lower()
                
                # Check for 72239 extrusion subassembly
                if "72239-xxx" in path:
                    trans = comp.Transform2
                    if trans:
                        data = trans.ArrayData
                        x_pos = data[9]
                        y_pos = data[10]
                        is_horizontal = abs(y_pos) > abs(x_pos)
                    else:
                        is_horizontal = True
                        
                    is_n = "72239-xxx-36.00-n" in path
                    
                    if is_horizontal:
                        replacement_path = out_asm_w
                    else:
                        if is_n:
                            replacement_path = out_asm_hn
                        else:
                            replacement_path = out_asm_h
                            
                    log.info(f"Replacing frame instance {name} with {os.path.basename(replacement_path)}...")
                    swMirror.ClearSelection2(True)
                    comp.Select2(False, 0)
                    success = swAssy.ReplaceComponents2(replacement_path, "Default", False, 2, True)
                    log.info(f"  ReplaceComponents2 status: {success}")

                # Check for LED subassembly
                elif "82000" in path or "82180" in path:
                    log.info(f"Replacing LED instance {name} with {os.path.basename(out_led_asm)}...")
                    swMirror.ClearSelection2(True)
                    comp.Select2(False, 0)
                    success = swAssy.ReplaceComponents2(out_led_asm, "Default", False, 2, True)
                    log.info(f"  ReplaceComponents2 status: {success}")

            # Set mirror properties
            mirror_props = {
                "PartNumber": cpn + "-M",
                "Width": f"{width:.2f}",
                "Height": f"{height:.2f}",
            }
            self._set_custom_properties(swMirror, mirror_props)

            # Traverse and configure/suppress/delete option-specific components in mirror
            button_needed = inp.Ava or inp.Keen or inp.Vive
            to_delete = []
            clock_comps = []
            defogger_comps = []
            
            # Re-read components after replacements
            components = swMirror.GetComponents(True)
            for comp in components:
                c_path = comp.GetPathName.lower()
                c_name = comp.Name2
                
                if "js3-ava-keen-vive-1-button-assembly" in c_path:
                    if not button_needed:
                        to_delete.append(comp)
                    else:
                        comp.SetSuppression2(2)  # 2 = FullyResolved
                        button_config = None
                        if inp.Ava:
                            button_config = f"64167-{inp.DimmingType}"
                        elif inp.Keen:
                            if result.DriverQty < 2:
                                button_config = f"61857-{inp.DimmingType}"
                            else:
                                button_config = f"61857-{inp.DimmingType}-2"
                        elif inp.Vive:
                            button_config = "83047"
                            
                        if button_config:
                            comp.ReferencedConfiguration = button_config
                            log.info(f"Resolved button assembly and set configuration to: {button_config}")

                elif "clock option-mirror" in c_path or "clock option-mirror" in c_name.lower():
                    clock_comps.append(comp)
                    
                elif "defogger-js" in c_path or "defogger-js" in c_name.lower():
                    defogger_comps.append(comp)

            # Assert Clock component state
            clock_path = str(Path(vault) / "COMPONENTS" / "STANDARD PARTS" / "CLOCK OPTION" / "CLOCK OPTION-MIRROR.sldasm")
            if inp.Clock:
                clock_config = result.ClockConfig
                if clock_comps:
                    for comp in clock_comps:
                        comp.SetSuppression2(2)
                        comp.ReferencedConfiguration = clock_config
                        log.info(f"Configured existing clock component: {comp.Name2} with {clock_config}")
                else:
                    log.info("Adding missing clock component programmatically...")
                    try:
                        self.swApp.OpenDoc6(clock_path, 2, 1, "", errors, warnings)
                        comp = swMirror.AddComponent4(clock_path, clock_config, 0.0, 0.0, 0.0)
                        if comp:
                            log.info(f"Successfully added clock: {comp.Name2} | Config: {comp.ReferencedConfiguration}")
                    except Exception as e:
                        log.error(f"Error adding clock component: {e}")
            else:
                for comp in clock_comps:
                    to_delete.append(comp)

            # Assert Defogger components state
            defogger_path = str(Path(vault) / "COMPONENTS" / "STANDARD PARTS" / "Configurator Parts" / "DEFOGGER-JS.SLDPRT")
            if inp.Defogger:
                defogger_config = result.DefoggerConfig
                defogger_qty = result.DefoggerQty
                
                for comp in defogger_comps[:defogger_qty]:
                    comp.SetSuppression2(2)
                    comp.ReferencedConfiguration = defogger_config
                    log.info(f"Configured existing defogger component: {comp.Name2} with {defogger_config}")
                    
                if len(defogger_comps) > defogger_qty:
                    for comp in defogger_comps[defogger_qty:]:
                        to_delete.append(comp)
                        
                curr_count = len(defogger_comps)
                if curr_count < defogger_qty:
                    self.swApp.OpenDoc6(defogger_path, 1, 1, "", errors, warnings)
                while curr_count < defogger_qty:
                    log.info(f"Adding missing defogger component {curr_count + 1} of {defogger_qty}...")
                    try:
                        comp = swMirror.AddComponent4(defogger_path, defogger_config, 0.0, 0.0, 0.0)
                        if comp:
                            log.info(f"Successfully added defogger: {comp.Name2} | Config: {comp.ReferencedConfiguration}")
                    except Exception as e:
                        log.error(f"Error adding defogger component: {e}")
                    curr_count += 1
            else:
                for comp in defogger_comps:
                    to_delete.append(comp)

            # Execute physical deletion of unneeded option components
            if to_delete:
                for comp in to_delete:
                    comp.Select2(False, 0)
                    swMirror.EditDelete()
                    log.info(f"Physically deleted component from mirror: {comp.Name2}")

            swMirror.ForceRebuild3(False)
            swMirror.Save3(swSaveAsOptions_Silent, errors, warnings)
            self.swApp.CloseDoc(swMirror.GetTitle)
            log.info("Mirror assembly configured and saved.")
        else:
            raise RuntimeError(f"Failed to open mirror assembly copy: {out_mirror}")

        # B. Open and configure Chassis Assembly
        log.info(f"Opening Chassis copy: {out_chassis}")
        swChassis = self.swApp.OpenDoc6(str(out_chassis), swDocASSEMBLY, 1, "", errors, warnings)
        if swChassis:
            self.swApp.ActivateDoc3(swChassis.GetTitle, True, 2, errors)
            
            # Read coordinates from Chassis assembly components
            components = swChassis.GetComponents(True)
            for comp in components:
                if comp.IsSuppressed:
                    continue
                c_name = comp.Name2.upper()
                if "81330" in c_name or "DRIVER-MODULE" in c_name:
                    trans = comp.Transform2
                    if trans:
                        powerBoxY = trans.ArrayData[10] * 39.3701
                        log.info(f"Found active Driver Module Y position relative to origin: {powerBoxY:.4f} in")
                elif "HANGER" in c_name or "81014" in c_name:
                    trans = comp.Transform2
                    if trans:
                        hangerY = trans.ArrayData[10] * 39.3701
                        log.info(f"Found active Hanger Bracket Y position relative to origin: {hangerY:.4f} in")

            # Set driver configuration
            driver_found = False
            for comp in components:
                name = comp.Name2
                if "81330-DRIVER-MODULE" in name:
                    log.info(f"Found driver component: {name}, resolving and setting configuration to Default")
                    comp.SetSuppression2(2)  # FullyResolved
                    comp.ReferencedConfiguration = "Default"
                    driver_found = True
                    break
            if not driver_found:
                log.warning("Driver module component not found in Chassis assembly!")
            else:
                swChassis.ForceRebuild3(False)

            # Traverse recursively to configure chassis clock & defogger wires
            all_chassis_comps = swChassis.GetComponents(False)
            if all_chassis_comps:
                for comp in all_chassis_comps:
                    comp_path = comp.GetPathName.lower()
                    comp_name = comp.Name2.lower()
                    
                    if "clock option-chassis" in comp_path or "clock option-chassis" in comp_name:
                        if inp.Clock and inp.ClockType in ["CK1", "CK2"]:
                            comp.SetSuppression2(2)
                            clock_wire_config = f"{inp.ClockType} SHORT WIRES"
                            comp.ReferencedConfiguration = clock_wire_config
                            log.info(f"Unsuppressed clock wire harness: {comp_name} and set config to: {clock_wire_config}")
                        else:
                            comp.SetSuppression2(0)
                            log.info(f"Suppressed clock wire harness component in chassis.")
                            
                    elif "74826-harness-wire-defogger-driver-box" in comp_path or "74826-harness-wire-defogger-driver-box" in comp_name:
                        if inp.Defogger:
                            comp.SetSuppression2(2)
                            log.info(f"Unsuppressed defogger wire harness in chassis driver box.")
                        else:
                            comp.SetSuppression2(0)
                            log.info(f"Suppressed defogger wire harness in chassis driver box.")

            chassis_props = {
                "PartNumber": cpn + "-C",
                "Width": f"{width:.2f}",
                "Height": f"{height:.2f}",
                "Power_Required_W": f"{result.PowerRequirement:.2f}",
                "Power_Available_W": str(result.DriverWattage),
                "Driver_Qty": str(result.DriverQty),
                "Finish": inp.Finish,
                "LED_Color_Temp": inp.LEDColorTemp,
                "Lighting": inp.Lighting,
                "MountType": inp.MountType,
                "Voltage": inp.Voltage,
                "Driver_Type": result.DriverType,
                "Enclosure_PN": result.DriverEnclosurePN,
                "LED_PN": result.LEDPN,
                "LED_Cut_Length_In": f"{result.LEDCutLengthIn:.2f}",
                "LED_Segments": str(result.LEDCuttableSegmentsFinal),
                "LED_Strips": str(result.LEDStripsRequiredEnc),
            }
            self._set_custom_properties(swChassis, chassis_props)

            # Update SA sketch dimensions for layout and mates
            chassis_dims = {
                "D1@SA": width * 0.0254,
                "D7@SA": height * 0.0254,
            }
            for name, val in chassis_dims.items():
                dim = swChassis.Parameter(name)
                if dim:
                    dim.SystemValue = val
                    log.info(f"Updated chassis dimension {name} to {val / 0.0254:.4f} in")
                else:
                    log.warning(f"Chassis dimension {name} not found in assembly.")

            swChassis.ForceRebuild3(False)
            swChassis.Save3(swSaveAsOptions_Silent, errors, warnings)
            self.swApp.CloseDoc(swChassis.GetTitle)
            log.info("Chassis assembly configured and saved.")
        else:
            raise RuntimeError(f"Failed to open chassis assembly copy: {out_chassis}")

        # C. Configure Drawing and export PDF
        log.info(f"Opening Drawing copy: {out_drawing}")
        swDrawing = self.swApp.OpenDoc6(str(out_drawing), swDocDRAWING, 1, "", errors, warnings)
        if swDrawing:
            # Re-scale sheets without changing the sheet format
            opts = parse_cpn_options(cpn)
            parsed_notes = _get_drawing_notes_from_opts(width, height, opts)
            num_notes = len(parsed_notes)
            
            p1_scale = self._calculate_page_1_scale(width, height, num_notes)
            p2_scale = self._calculate_page_2_scale(width, height, max(1, num_notes - 2))
            log.info(f"Applying scales Page 1 (1:{p1_scale}), Page 2 (1:{p2_scale}) to drawing sheets.")
            
            sheet_names = swDrawing.GetSheetNames
            for i, sheet_name in enumerate(sheet_names):
                if i >= 2:
                    break
                swDrawing.ActivateSheet(sheet_name)
                scale_den = p1_scale if i == 0 else p2_scale
                
                # Apply SetProperties on the sheet instead of SetupSheet6
                sheet = swDrawing.GetCurrentSheet
                if sheet:
                    props = sheet.GetProperties2
                    if props:
                        ret = sheet.SetProperties(
                            int(props[0]),      # paperSize
                            int(props[1]),      # templateType
                            1.0,                # scale1 (numerator)
                            float(scale_den),   # scale2 (denominator)
                            bool(props[4]),     # firstPage
                            float(props[5]),    # width
                            float(props[6])     # height
                        )
                        log.info(f"Updated properties for sheet '{sheet_name}' (Scale 1:{scale_den}). Success: {ret}")
                    else:
                        log.error(f"Failed to get properties for sheet '{sheet_name}'")
                else:
                    log.error(f"Failed to get sheet object for sheet '{sheet_name}'")
                
                # Traverse views on this sheet and force them to use the sheet scale
                view = swDrawing.GetFirstView
                if view:
                    view = view.GetNextView  # Skip the sheet view itself
                    while view:
                        try:
                            view.UseSheetScale = True
                            log.info(f"Set view.UseSheetScale = True on view '{view.Name}' on sheet '{sheet_name}'")
                        except Exception as e:
                            log.error(f"Failed to set UseSheetScale on view '{view.Name}' on sheet '{sheet_name}': {e}")
                        view = view.GetNextView
            
            # Reposition Section Line A-A on Sheet 1
            try:
                self._reposition_section_line_aa(swDrawing)
            except Exception as e:
                log.error(f"Failed to reposition Section Line A-A: {e}")
            
            # Update Sketch34 dimensioning on Sheet 2
            try:
                self._update_sketch34_dimensions(swDrawing, width, height, powerBoxY, hangerY)
            except Exception as e:
                log.error(f"Failed to update Sketch34 dimensions: {e}")
            FINISH_MAP = {
                "BK": "MATTE BLACK", "BK05": "MATTE BLACK",
                "NK": "ETCHED NICKEL", "NK04": "ETCHED NICKEL",
                "CH": "ETCHED CHROME", "CH04": "ETCHED CHROME", "CH11": "BRIGHT CHROME",
                "BN": "BRUSHED NICKEL", "BN04": "BRUSHED NICKEL",
                "BR": "BRUSHED BRASS", "BR02": "BRUSHED BRASS", "BR21": "BRIGHT BRASS",
                "BZ24": "ETCHED GOLDEN BRONZE", "BZ47": "BRONZE",
                "GU": "GUN METAL", "GU06": "GUN METAL",
                "WH": "GLOSS WHITE", "WH01": "GLOSS WHITE",
            }
            finish_name = FINISH_MAP.get(inp.Finish, inp.Finish)
            try:
                cct_val = int(inp.LEDColorTemp.replace('K', '')) * 100
                cct_str = f"{cct_val:,}"
            except Exception:
                cct_str = inp.LEDColorTemp

            lm_ft = 302.0
            total_lumens = round(result.LEDCutLengthIn * lm_ft / 12.0)
            
            drawing_props = {
                "Sales Aid": f"RAD4-{width:.2f}X{height:.2f}-{inp.Finish}-{inp.LEDColorTemp}",
                "Frame Finish": finish_name,
                "ColorTemp": cct_str,
                "LEDLength": str(round(result.LEDCutLengthIn)),
                "Wattage": str(round(result.PowerRequirement)),
                "PowerRequirements": f"120 OR 277 VOLTS, {result.PowerRequirement / 120.0 * 0.9:.2f} OR {result.PowerRequirement / 277.0 * 0.9:.2f} AMPS",
                "Lumens": f"{total_lumens:,} @ 302 LM/FT",
                "PartNumber": cpn,
                "Description": f"RAD4 {width:.2f}X{height:.2f} {finish_name}",
            }
            self._set_custom_properties(swDrawing, drawing_props)
            log.info("Drawing custom properties set successfully.")

            # Option A dynamic notes update and suppression
            try:
                self._update_drawing_notes_option_a(swDrawing, inp, result)
            except Exception as e:
                log.error(f"Failed to update drawing notes via Option A: {e}")

            # Rebuild twice to ensure all properties propagate to title blocks
            swDrawing.ForceRebuild3(False)
            swDrawing.ForceRebuild3(False)

            # Export to PDF in target output directory
            os.makedirs(output_dir, exist_ok=True)
            export_data = self.swApp.GetExportFileData(1) # 1 = swExportPdfData
            success_pdf = swDrawing.Extension.SaveAs(pdf_path, 0, swSaveAsOptions_Silent, export_data, errors, warnings)
            log.info(f"PDF Export status: {success_pdf} to {pdf_path}")

            swDrawing.Save3(swSaveAsOptions_Silent, errors, warnings)
            self.swApp.CloseDoc(swDrawing.GetTitle)
            log.info("Drawing configured and saved.")
        else:
            raise RuntimeError(f"Failed to open drawing copy: {out_drawing}")

        # 7. Export BOM to Excel
        self._export_bom(inp, result, bom_path, bom_pdf_path)

        return {
            "chassis_assembly": str(out_chassis),
            "mirror_assembly": str(out_mirror),
            "sales_aid_drawing": str(out_drawing),
            "sales_aid_pdf": pdf_path,
            "bom_excel": bom_path,
            "bom_pdf": bom_pdf_path
        }

    def _configure_local_extrusions(self, vault: str, output_dir: str, width: float, height: float) -> tuple[str, str, str]:
        errors = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        warnings = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

        comp_72_dir = Path(vault) / "Products" / "JS3" / "COMPONENTS" / "72000"
        comp_64_dir = Path(vault) / "COMPONENTS" / "64000"
        comp_64_js3_dir = Path(vault) / "Products" / "JS3" / "COMPONENTS" / "64000"

        # Ensure Vault directories exist programmatically
        comp_72_dir.mkdir(parents=True, exist_ok=True)
        comp_64_js3_dir.mkdir(parents=True, exist_ok=True)

        # Width assembly & parts
        out_alu_w = comp_72_dir / f"72239-{width:.2f}-EXTRUSION-ALUMINUM.SLDPRT"
        out_dif_w = comp_64_js3_dir / f"64792-{width:.2f}-EXTRUSION-DIFFUSER.SLDPRT"
        out_asm_w = comp_72_dir / f"72239-XXX-{width:.2f}.SLDASM"

        temp_alu_w = comp_72_dir / "72239-36.00-EXTRUSION-ALUMINUM.SLDPRT"
        temp_dif_w = comp_64_js3_dir / "64792-36.00-EXTRUSION-DIFFUSER.SLDPRT"
        if not temp_dif_w.exists():
            temp_dif_w = comp_64_dir / "64792-36.00-EXTRUSION-DIFFUSER.SLDPRT"
        temp_asm_w = comp_72_dir / "72239-XXX-36.00.SLDASM"

        if not out_asm_w.exists() or not out_alu_w.exists():
            _copy_file_writable(temp_alu_w, out_alu_w)
            _copy_file_writable(temp_dif_w, out_dif_w)
            _copy_file_writable(temp_asm_w, out_asm_w)

            self.swApp.ReplaceReferencedDocument(str(out_asm_w), str(temp_alu_w), str(out_alu_w))
            self.swApp.ReplaceReferencedDocument(str(out_asm_w), str(temp_dif_w), str(out_dif_w))

            swPart = self.swApp.OpenDoc6(str(out_alu_w), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@Extrusion").SystemValue = width * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swPart = self.swApp.OpenDoc6(str(out_dif_w), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@Boss-Extrude1").SystemValue = width * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swAsm = self.swApp.OpenDoc6(str(out_asm_w), swDocASSEMBLY, 1, "", errors, warnings)
            if swAsm:
                swAsm.ForceRebuild3(False)
                swAsm.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swAsm.GetTitle)

        # Height assembly & parts
        out_alu_h = comp_72_dir / f"72239-{height:.2f}-EXTRUSION-ALUMINUM.SLDPRT"
        out_dif_h = comp_64_js3_dir / f"64792-{height:.2f}-EXTRUSION-DIFFUSER.SLDPRT"
        out_asm_h = comp_72_dir / f"72239-XXX-{height:.2f}.SLDASM"

        if not out_asm_h.exists() or not out_alu_h.exists():
            _copy_file_writable(temp_alu_w, out_alu_h)
            _copy_file_writable(temp_dif_w, out_dif_h)
            _copy_file_writable(temp_asm_w, out_asm_h)

            self.swApp.ReplaceReferencedDocument(str(out_asm_h), str(temp_alu_w), str(out_alu_h))
            self.swApp.ReplaceReferencedDocument(str(out_asm_h), str(temp_dif_w), str(out_dif_h))

            swPart = self.swApp.OpenDoc6(str(out_alu_h), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@Extrusion").SystemValue = height * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swPart = self.swApp.OpenDoc6(str(out_dif_h), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@Boss-Extrude1").SystemValue = height * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swAsm = self.swApp.OpenDoc6(str(out_asm_h), swDocASSEMBLY, 1, "", errors, warnings)
            if swAsm:
                swAsm.ForceRebuild3(False)
                swAsm.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swAsm.GetTitle)

        # Height N assembly & parts
        out_alu_hn = comp_72_dir / f"72239-{height:.2f}-N-EXTRUSION-ALUMINUM.SLDPRT"
        out_asm_hn = comp_72_dir / f"72239-XXX-{height:.2f}-N.SLDASM"

        temp_alu_hn = comp_72_dir / "72239-36.00-N-EXTRUSION-ALUMINUM.SLDPRT"
        temp_asm_hn = comp_72_dir / "72239-XXX-36.00-N.SLDASM"

        if not out_asm_hn.exists() or not out_alu_hn.exists():
            _copy_file_writable(temp_alu_hn, out_alu_hn)
            _copy_file_writable(temp_asm_hn, out_asm_hn)

            self.swApp.ReplaceReferencedDocument(str(out_asm_hn), str(temp_alu_hn), str(out_alu_hn))
            self.swApp.ReplaceReferencedDocument(str(out_asm_hn), str(temp_dif_w), str(out_dif_h))

            swPart = self.swApp.OpenDoc6(str(out_alu_hn), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@Extrusion").SystemValue = height * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swAsm = self.swApp.OpenDoc6(str(out_asm_hn), swDocASSEMBLY, 1, "", errors, warnings)
            if swAsm:
                swAsm.ForceRebuild3(False)
                swAsm.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swAsm.GetTitle)

        return str(out_asm_w), str(out_asm_h), str(out_asm_hn)

    def _configure_local_led(self, vault: str, output_dir: str, width: float, height: float, target_led_pn: str) -> str:
        errors = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        warnings = win32.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

        comp_82_dir = Path(vault) / "Products" / "JS3" / "COMPONENTS" / "82000"
        comp_led_parts_dir = Path(vault) / "Products" / "JS3" / "COMPONENTS" / "RAD4 LEDs"

        # Ensure Vault directories exist programmatically
        comp_82_dir.mkdir(parents=True, exist_ok=True)
        comp_led_parts_dir.mkdir(parents=True, exist_ok=True)

        # Compute physical perimeter to determine standard round length
        total_perimeter_mm = 2 * (width + height) * 25.4 - 49.58223168
        new_len = int(round(total_perimeter_mm / 50.0) * 50.0)

        out_led_part = comp_led_parts_dir / f"RAD3-LED-{new_len}-{width:.2f}X{height:.2f}.SLDPRT"
        out_led_asm = comp_82_dir / f"82180-RAD3-{new_len}MM-{width:.2f}X{height:.2f}.SLDASM"

        # Templates
        temp_led_part = comp_led_parts_dir / "RAD3-LED-3600-36.00X36.00.SLDPRT"
        if not temp_led_part.exists():
            temp_led_part = Path(vault) / "COMPONENTS" / "STANDARD PARTS" / "Configurator Parts" / "RAD3" / "LEDs" / "RAD3-LED-3600-36.00X36.00.SLDPRT"
        temp_led_asm = comp_82_dir / "82180-RAD3-3600MM-36.00X36.00.SLDASM"

        if not out_led_asm.exists() or not out_led_part.exists():
            _copy_file_writable(temp_led_part, out_led_part)
            _copy_file_writable(temp_led_asm, out_led_asm)

            self.swApp.ReplaceReferencedDocument(str(out_led_asm), str(temp_led_part), str(out_led_part))

            swPart = self.swApp.OpenDoc6(str(out_led_part), swDocPART, 1, "", errors, warnings)
            if swPart:
                swPart.Parameter("D1@MASTER SKETCH").SystemValue = width * 0.0254
                swPart.Parameter("D2@MASTER SKETCH").SystemValue = height * 0.0254
                swPart.ForceRebuild3(False)
                swPart.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swPart.GetTitle)

            swAsm = self.swApp.OpenDoc6(str(out_led_asm), swDocASSEMBLY, 1, "", errors, warnings)
            if swAsm:
                swAsm.ForceRebuild3(False)
                swAsm.Save3(swSaveAsOptions_Silent, errors, warnings)
                self.swApp.CloseDoc(swAsm.GetTitle)

        return str(out_led_asm)

    def _configure_custom_extrusions(self, vault: str, temp_w: float, temp_h: float, width: float, height: float):
        """No-op as custom extrusions are now handled programmatically via local copy-and-replace in generate."""
        pass

    def _set_custom_properties(self, swModel, props: dict):
        mgr = swModel.Extension.CustomPropertyManager("")
        for key, val in props.items():
            mgr.Add3(key, 30, str(val), 1)
            mgr.Set2(key, str(val))

    def _export_bom(self, inp: RAD4Inputs, result: RAD4Result, output_xlsx_path: str, output_pdf_path: str):
        try:
            import openpyxl
            import datetime
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            log.warning("openpyxl or datetime not installed — skipping Excel BOM.")
            return

        # Load generic BOM template from repository configurator/templates/
        templates_dir = Path(__file__).parent / "templates"
        template_bom = templates_dir / "RAD4-GENERIC-M-BOM.xlsx"
        
        if not template_bom.exists():
            log.error(f"Generic BOM template not found at {template_bom}")
            return

        try:
            log.info(f"Using generic BOM template: {template_bom}")
            _copy_file_writable(template_bom, Path(output_xlsx_path))
            
            # Open, update cells, and save to local output path only
            wb = openpyxl.load_workbook(output_xlsx_path)
            ws = wb.active
            
            width = inp.UnitWidth
            height = inp.UnitHeight
            
            # 1. Update Title Block Columns (Column D, which is 4)
            ws.cell(row=1, column=4, value=f"{result.CPN}-M")
            ws.cell(row=2, column=4, value=datetime.datetime.now().strftime("%m/%d/%Y"))
            ws.cell(row=4, column=4, value=f"RAD4-{width:.2f}X{height:.2f}-M-SHOP")
            ws.cell(row=5, column=4, value='=LEFT(CPN, FIND("-M", CPN)-1)&"-M-PLOT"')
            ws.cell(row=6, column=4, value="AI")
            ws.cell(row=7, column=4, value="AI")
            ws.cell(row=8, column=4, value=datetime.datetime.now().strftime("%m/%d/%Y"))
            
            # 2. Append option rows at the end of the table
            row_idx = 30
            while True:
                val = ws.cell(row_idx, 1).value
                if val is None:
                    break
                row_idx += 1
                
            if inp.Defogger:
                ws.insert_rows(row_idx, 1)
                for col in range(1, 10):
                    orig_cell = ws.cell(row_idx - 1, col)
                    new_cell = ws.cell(row_idx, col)
                    if orig_cell.font:
                        new_cell.font = openpyxl.styles.Font(name=orig_cell.font.name, size=orig_cell.font.size, bold=orig_cell.font.bold, italic=orig_cell.font.italic)
                    if orig_cell.alignment:
                        new_cell.alignment = openpyxl.styles.Alignment(horizontal=orig_cell.alignment.horizontal, vertical=orig_cell.alignment.vertical)
                    if orig_cell.border:
                        new_cell.border = openpyxl.styles.Border(left=orig_cell.border.left, right=orig_cell.border.right, top=orig_cell.border.top, bottom=orig_cell.border.bottom)
                
                df_base = result.DefoggerConfig.split('-')[0]
                ws.cell(row_idx, 1, value=99)
                ws.cell(row_idx, 2, value=f"{df_base}-1")
                ws.cell(row_idx, 4, value="DEFOGGER")
                ws.cell(row_idx, 9, value=result.DefoggerQty)
                row_idx += 1

            # Append Driver Module
            driver_pn = result.DriverEnclosurePN or result.BOM.get('Driver_Enclosure', '')
            if driver_pn:
                ws.insert_rows(row_idx, 1)
                for col in range(1, 10):
                    orig_cell = ws.cell(row_idx - 1, col)
                    new_cell = ws.cell(row_idx, col)
                    if orig_cell.font:
                        new_cell.font = openpyxl.styles.Font(name=orig_cell.font.name, size=orig_cell.font.size, bold=orig_cell.font.bold, italic=orig_cell.font.italic)
                    if orig_cell.alignment:
                        new_cell.alignment = openpyxl.styles.Alignment(horizontal=orig_cell.alignment.horizontal, vertical=orig_cell.alignment.vertical)
                    if orig_cell.border:
                        new_cell.border = openpyxl.styles.Border(left=orig_cell.border.left, right=orig_cell.border.right, top=orig_cell.border.top, bottom=orig_cell.border.bottom)
                
                ws.cell(row_idx, 1, value=99)
                ws.cell(row_idx, 2, value=driver_pn)
                ws.cell(row_idx, 4, value="DRIVER MODULE")
                ws.cell(row_idx, 9, value=1)
                row_idx += 1

            # Append Standoff Kit
            standoff_pn = result.BOM.get('Standoff_Kit', '')
            if standoff_pn:
                ws.insert_rows(row_idx, 1)
                for col in range(1, 10):
                    orig_cell = ws.cell(row_idx - 1, col)
                    new_cell = ws.cell(row_idx, col)
                    if orig_cell.font:
                        new_cell.font = openpyxl.styles.Font(name=orig_cell.font.name, size=orig_cell.font.size, bold=orig_cell.font.bold, italic=orig_cell.font.italic)
                    if orig_cell.alignment:
                        new_cell.alignment = openpyxl.styles.Alignment(horizontal=orig_cell.alignment.horizontal, vertical=orig_cell.alignment.vertical)
                    if orig_cell.border:
                        new_cell.border = openpyxl.styles.Border(left=orig_cell.border.left, right=orig_cell.border.right, top=orig_cell.border.top, bottom=orig_cell.border.bottom)
                
                ws.cell(row_idx, 1, value=99)
                ws.cell(row_idx, 2, value=standoff_pn)
                ws.cell(row_idx, 4, value="STANDOFF KIT")
                ws.cell(row_idx, 9, value=1)
                row_idx += 1


            # 3. Clean up empty/inactive option rows in the table
            r = 30
            while r < row_idx:
                item_no = ws.cell(r, 1).value
                if item_no is None:
                    break
                    
                pn_val = str(ws.cell(r, 2).value or "").strip()
                
                # Check if the row should be deleted because of inactive option
                delete_row = False
                if "SEARCH({\"CK\"}" in pn_val and not inp.Clock:
                    delete_row = True
                elif "SEARCH({\"KC\"" in pn_val and not (inp.Ava or inp.Keen or inp.Vive):
                    delete_row = True
                elif "OR(Width >= 60" in pn_val and max(width, height) < 60:
                    delete_row = True
                    
                if delete_row:
                    ws.delete_rows(r, 1)
                    row_idx -= 1
                else:
                    r += 1

            # 4. Renumber item numbers sequentially in Column A
            curr_item = 1
            for r_num in range(30, r):
                ws.cell(r_num, 1, value=curr_item)
                curr_item += 1

            wb.save(output_xlsx_path)
            log.info(f"BOM exported with generic template to {output_xlsx_path}")
            self._export_bom_to_pdf(output_xlsx_path, output_pdf_path)
            return
        except Exception as e:
            log.error(f"Error customizing generic BOM: {e}")

        # Fallback to basic BOM generation from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"

        header_fill = PatternFill("solid", fgColor="1F3864")
        
        ws.cell(row=1, column=1, value=f"BOM — {result.CPN}")
        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=1, column=1).fill = header_fill

        headers = ["BOM Item", "Part Number", "Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")

        row = 3
        for item, pn in result.BOM.items():
            if pn:
                ws.cell(row=row, column=1, value=item.replace("_", " "))
                ws.cell(row=row, column=2, value=pn)
                row += 1

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 40

        # Save to local output target only
        wb.save(output_xlsx_path)
        log.info(f"BOM exported (fallback) to: {output_xlsx_path}")
        self._export_bom_to_pdf(output_xlsx_path, output_pdf_path)

    def _export_bom_to_pdf(self, bom_xlsx_path: str, output_pdf_path: str):
        try:
            import win32com.client
            import pythoncom
            log.info(f"Converting Excel BOM to PDF: {bom_xlsx_path}")
            
            # Initialize COM and dispatch Excel
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            abs_xlsx = os.path.abspath(bom_xlsx_path)
            abs_pdf = os.path.abspath(output_pdf_path)
            
            wb = excel.Workbooks.Open(abs_xlsx)
            wb.ExportAsFixedFormat(0, abs_pdf)
            wb.Close(False)
            excel.Quit()
            log.info(f"Successfully exported BOM PDF to: {abs_pdf}")
        except Exception as e:
            log.error(f"Failed to export BOM PDF via Excel COM API: {e}")

    def _parse_defogger_config(self, config_name: str) -> tuple[str, str]:
        c_upper = config_name.upper()
        # Extract size from config name e.g. "15230-120V-22.5x10.5" -> "22.5X10.5"
        m = re.search(r'-(\d+\.?\d*X\d+\.?\d*)$', c_upper)
        size = m.group(1) if m else "10.5X10.5"
        
        if "15229" in c_upper:
            watts = "15"
        elif "15230" in c_upper:
            watts = "25"
        elif "15231" in c_upper:
            watts = "50"
        elif "15232" in c_upper:
            watts = "100"
        else:
            watts = "25"
            
        return size, watts

    def _update_drawing_notes_option_a(self, swDrawing, inp: RAD4Inputs, result: RAD4Result):
        log.info("Updating drawing notes using Option A dynamic rules...")
        
        # Calculate dynamic text blocks
        amp_120 = (result.PowerRequirement / 120.0) * 0.9
        amp_277 = (result.PowerRequirement / 277.0) * 0.9
        cct_map = {
            "27K": "2,700",
            "30K": "3,000",
            "35K": "3,500",
            "40K": "4,000"
        }
        cct_val = cct_map.get(inp.LEDColorTemp, "3,000")
        lm_ft = 302.0
        total_lumens = round(result.LEDCutLengthIn * lm_ft / 12.0)
        
        opts = parse_cpn_options(result.CPN)
        parsed_notes = _get_drawing_notes_from_opts(inp.UnitWidth, inp.UnitHeight, opts)

        # Calculate dynamic page scales for the scale notes
        num_notes = len(parsed_notes)
        p1_scale = self._calculate_page_1_scale(inp.UnitWidth, inp.UnitHeight, num_notes)
        p2_scale = self._calculate_page_2_scale(inp.UnitWidth, inp.UnitHeight, max(1, num_notes - 2))

        # 1. Main specs block
        if inp.Voltage == "277V":
            power_str = f"POWER REQUIREMENTS:\n277 VOLTS, {amp_277:.2f} AMPS"
        else:
            power_str = f"POWER REQUIREMENTS:\n120 VOLTS, {amp_120:.2f} AMPS"
            
        spec_text = (
            f"{power_str}\n\n"
            f"LED SPECIFICATION:\n"
            f"LED TYPE: REPLACEABLE FLEX STRIP\n"
            f"LENGTH (IN): {round(result.LEDCutLengthIn)}\n"
            f"WATTAGE (W): {round(result.PowerRequirement)}\n"
            f"CALCULATED L70 LIFESPAN (HRS): 140,000\n"
            f"CCT(K): {cct_val}\n"
            f"TOTAL INITIAL LUMENS PER FIXTURE: {total_lumens:,} @ 302 LM/FT\n"
            f"CRI: 90+"
        )
        
        if "DEFOGGER" in parsed_notes:
            pad_size, _ = self._parse_defogger_config(result.DefoggerConfig)
            wattage_str = parsed_notes["DEFOGGER"].split(":")[-1].strip()
            spec_text += (
                f"\n\nDEFOGGER SPECIFICATION:\n"
                f"WATTAGE: {wattage_str.split('@')[0].strip()}\n"
                f"VOLTAGE: {wattage_str.split('@')[1].strip()}"
            )
            
        spec_text += (
            f"\n\nFIXTURE SPECIFICATION:\n"
            f"WEIGHT: $PRPSHEET:\"Weight\" LBS (EST.)"
        )
        
        # 2. Spec instructions block
        spec_inst = "SPECIFICATION:\n" + parsed_notes.get("SPECIFICATION", "")
        if "ATTENTION" in parsed_notes:
            spec_inst += "\n\nATTENTION:\n" + parsed_notes["ATTENTION"]
        if "DEFOGGER DISCLAIMER (KEEN)" in parsed_notes:
            spec_inst += "\n\nKEEN / DEFOGGER DISCLAIMER:\n" + parsed_notes["DEFOGGER DISCLAIMER (KEEN)"]
            
        # 3. Dimming compatibility
        dimmer_text = parsed_notes.get("DIMMING", "")
        if dimmer_text:
            dimmer_text = "DIMMER COMPATIBILITY:\n" + dimmer_text
            
        # 4. Defogger detail
        defogger_text = ""
        outline_text = ""
        if "DEFOGGER" in parsed_notes:
            pad_size, _ = self._parse_defogger_config(result.DefoggerConfig)
            wattage_str = parsed_notes["DEFOGGER"].split(":")[-1].strip()
            defogger_text = f"DEFOGGER SPECIFICATIONS:\n{pad_size}\n{wattage_str}"
            outline_text = f"OUTLINE OF\nDEFOGGER\n{pad_size}"
            
        # 5. Button Callout
        if inp.Keen:
            button_text = "KEEN 1-TOUCH CONTROL BUTTON\nSEE SPECIFICATION SHEET\nFOR ADDITIONAL DETAILS"
        elif inp.Ava:
            button_text = "AVA 1-TOUCH CONTROL BUTTON\nSEE SPECIFICATION SHEET\nFOR ADDITIONAL DETAILS"
        elif inp.Vive:
            button_text = "VIVE CONTROL BUTTON\nSEE SPECIFICATION SHEET\nFOR ADDITIONAL DETAILS"
        else:
            button_text = ""
            
        # Perform drawing sheet traversal
        try:
            sheet_names = swDrawing.GetSheetNames
            for idx, s_name in enumerate(sheet_names):
                swDrawing.ActivateSheet(s_name)
                current_scale = p1_scale if idx == 0 else p2_scale
                
                view = swDrawing.GetFirstView
                while view:
                    note = view.GetFirstNote
                    while note:
                        text = note.GetText
                        name = note.GetName
                        text_lower = text.lower() if text else ""
                        
                        # Apply rules based on text content and name matches
                        updated = False
                        
                        # Rule 0: Title block scale note
                        if name == "DetailItem371" or text_lower.startswith("scale:"):
                            note.SetText(f"SCALE: 1:{current_scale}")
                            log.info(f"Updated scale note '{name}' on sheet '{s_name}' to 'SCALE: 1:{current_scale}'")
                            updated = True
                            
                        # Rule 1: Main specs block
                        elif name == "DetailItem378" or ("power requirements:" in text_lower and "led specification:" in text_lower):
                            note.SetText(spec_text)
                            log.info(f"Updated main specs block '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 2: Spec instructions
                        elif name in ("DetailItem436", "DetailItem435", "DetailItem434") or ("specification:" in text_lower and "bring mc cable" in text_lower):
                            note.SetText(spec_inst)
                            log.info(f"Updated spec instructions '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 3: Dimmer compatibility
                        elif name == "DetailItem409" or name == "DetailItem432" or "dimmer compatibility:" in text_lower:
                            note.SetText(dimmer_text)
                            log.info(f"Updated dimmer compatibility '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 4: Defogger specifications box
                        elif name == "DetailItem415" or "defogger specifications:" in text_lower or (text_lower.strip().startswith("defogger:") and "wattage:" in text_lower):
                            note.SetText(defogger_text)
                            log.info(f"Updated defogger spec box '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 5: Defogger outline callout
                        elif name == "DetailItem392" or "outline of\ndefogger" in text_lower or "outline of\r\ndefogger" in text_lower:
                            note.SetText(outline_text)
                            log.info(f"Updated defogger outline '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 6: Touch button callout
                        elif name == "DetailItem437" or "1-touch control" in text_lower or "control button" in text_lower:
                            note.SetText(button_text)
                            log.info(f"Updated touch button callout '{name}' on sheet '{s_name}'")
                            updated = True
                            
                        # Rule 7: Clean up duplicates (clear standalone warning notes since they are now appended to spec_inst)
                        elif not updated:
                            if "earth ground in\naccordance with nec" in text_lower or "earth ground in\r\naccordance with nec" in text_lower:
                                note.SetText("")
                                log.info(f"Cleared standalone grounding warning '{name}' to avoid duplicate")
                            elif "keen / defogger disclaimer:" in text_lower:
                                note.SetText("")
                                log.info(f"Cleared standalone keen disclaimer '{name}' to avoid duplicate")
                                
                        note = note.GetNext
                    view = view.GetNextView
        except Exception as e:
            log.error(f"Error during Sheet/View note traversal: {e}")

    def _calculate_page_1_scale(self, width: float, height: float, num_notes: int) -> int:
        LADDER = [8, 10, 12, 14, 16, 18, 20, 24]
        SHEET_W = 17.0
        SHEET_H = 11.0
        TITLE_BLOCK_H = 1.25
        MARGIN = 0.40
        PAGE_SEP_H = 0.20
        NOTE_AVG_H = 1.50
        NOTES_COL_BASE_W = 2.80
        NOTES_COL_GROWTH_W = 0.30
        P1_SIDE_VIEW_W = 1.20
        P1_DIM_LEADER_W = 1.20
        P1_DIM_LEADER_H = 1.20
        CONVENTIONAL_TARGET_BASE = 4.25
        CONVENTIONAL_NOTES_DROP = 0.08
        CONVENTIONAL_WIDE_BONUS = 0.25
        CONVENTIONAL_SQUARE_DROP = 0.40
        CONVENTIONAL_VERTICAL_PB = 0.05
        GEOMETRY_SAFETY = 1.10

        pb = "horizontal" if width >= 22.0 else "vertical"
        
        # Notes column width
        notes_w = 0.0
        if num_notes > 0:
            notes_w = NOTES_COL_BASE_W + NOTES_COL_GROWTH_W * (num_notes - 1)
            col_h_avail = SHEET_H - TITLE_BLOCK_H - 2 * MARGIN
            needed_h = num_notes * NOTE_AVG_H
            if needed_h > col_h_avail:
                notes_w *= (needed_h / col_h_avail)
            notes_w = min(notes_w, 6.0)

        # Clear area
        clear_w = SHEET_W - 2 * MARGIN - notes_w - PAGE_SEP_H
        clear_h = SHEET_H - TITLE_BLOCK_H - 2 * MARGIN
        view_w = clear_w - P1_SIDE_VIEW_W - P1_DIM_LEADER_W
        view_h = clear_h - P1_DIM_LEADER_H

        n_required = max(width / max(0.5, view_w), height / max(0.5, view_h)) * GEOMETRY_SAFETY

        # Conventional target
        t = CONVENTIONAL_TARGET_BASE - CONVENTIONAL_NOTES_DROP * max(0, num_notes - 1)
        r = width / height if height > 0 else 1.0
        if r >= 1.2:
            t += CONVENTIONAL_WIDE_BONUS
        elif r <= 1/1.2:
            pass
        else:
            t -= CONVENTIONAL_SQUARE_DROP
        if pb == "vertical":
            t -= CONVENTIONAL_VERTICAL_PB
        target = max(2.5, t)
        
        n_preferred = max(width, height) / target
        
        raw_n = max(n_required, n_preferred)
        if raw_n <= 0:
            return LADDER[0]
        for s in LADDER:
            if s >= raw_n:
                return s
        return LADDER[-1]

    def _calculate_page_2_scale(self, width: float, height: float, num_notes_p2: int) -> int:
        LADDER = [8, 10, 12, 14, 16, 18, 20, 24]
        SHEET_W = 17.0
        SHEET_H = 11.0
        TITLE_BLOCK_H = 1.25
        MARGIN = 0.40
        PAGE_SEP_H = 0.20
        NOTE_AVG_H = 1.50
        NOTES_COL_BASE_W = 2.80
        NOTES_COL_GROWTH_W = 0.30
        P2_DIM_LEADER_W = 1.00
        P2_DIM_LEADER_H = 1.20
        DETAIL_A_HORIZ_W = 4.80
        DETAIL_A_HORIZ_H = 3.00
        DETAIL_A_VERT_W = 3.00
        DETAIL_A_VERT_H = 4.80
        CONVENTIONAL_TARGET_BASE = 4.25
        CONVENTIONAL_NOTES_DROP = 0.08
        CONVENTIONAL_WIDE_BONUS = 0.25
        CONVENTIONAL_SQUARE_DROP = 0.40
        CONVENTIONAL_VERTICAL_PB = 0.05
        P2_TARGET_BONUS = 0.50
        GEOMETRY_SAFETY = 1.10

        pb = "horizontal" if width >= 22.0 else "vertical"
        
        # Notes column width
        notes_w = 0.0
        if num_notes_p2 > 0:
            notes_w = NOTES_COL_BASE_W + NOTES_COL_GROWTH_W * (num_notes_p2 - 1)
            col_h_avail = SHEET_H - TITLE_BLOCK_H - 2 * MARGIN
            needed_h = num_notes_p2 * NOTE_AVG_H
            if needed_h > col_h_avail:
                notes_w *= (needed_h / col_h_avail)
            notes_w = min(notes_w, 6.0)

        # Clear area
        clear_w = SHEET_W - 2 * MARGIN - notes_w - PAGE_SEP_H
        clear_h = SHEET_H - TITLE_BLOCK_H - 2 * MARGIN
        if pb == "horizontal":
            det_w, det_h = DETAIL_A_HORIZ_W, DETAIL_A_HORIZ_H
        else:
            det_w, det_h = DETAIL_A_VERT_W, DETAIL_A_VERT_H
        view_w = clear_w - det_w - P2_DIM_LEADER_W
        view_h = clear_h - P2_DIM_LEADER_H

        n_required = max(width / max(0.5, view_w), height / max(0.5, view_h)) * GEOMETRY_SAFETY

        # Conventional target
        t = CONVENTIONAL_TARGET_BASE - CONVENTIONAL_NOTES_DROP * max(0, num_notes_p2 - 1)
        r = width / height if height > 0 else 1.0
        if r >= 1.2:
            t += CONVENTIONAL_WIDE_BONUS
        elif r <= 1/1.2:
            pass
        else:
            t -= CONVENTIONAL_SQUARE_DROP
        if pb == "vertical":
            t -= CONVENTIONAL_VERTICAL_PB
        target = max(2.5, t) + P2_TARGET_BONUS
        
        n_preferred = max(width, height) / target
        
        raw_n = max(n_required, n_preferred)
        if raw_n <= 0:
            return LADDER[0]
        for s in LADDER:
            if s >= raw_n:
                return s
        return LADDER[-1]

    def _update_sketch34_dimensions(self, swDrawing, W: float, H: float, powerBoxY: float, hangerY: float):
        log.info("Updating Sheet 2 Sketch34 mounting dimensions...")
        
        # Calculate mounting dimensions in inches
        hanger_w = W - 6.0 if W <= 48.0 else W - 8.0
        edge_inset = (W - hanger_w) / 2.0
        sec_spacing = 3.5 if W < 30.0 else 6.0
        
        pb_orient = "horizontal" if W >= 22.0 else "vertical"
        pb_width = 18.0 if pb_orient == "horizontal" else 10.5
        pb_height = 10.5 if pb_orient == "horizontal" else 18.0
        
        cable_h = H / 2.0 + powerBoxY
        pb_from_top = H / 2.0 - powerBoxY
        
        if abs(hangerY) > 0.01:
            hanger_h = H / 2.0 + hangerY
        else:
            hanger_h = H - 6.0
            
        dims = {
            "D4": W,
            "D5": H,
            "D7": hanger_w,
            "D10": edge_inset,
            "D8": sec_spacing,
            "D3": cable_h,
            "D6": pb_height,
            "D15": pb_width,
            "D11": pb_from_top,
            "D12": hanger_h,
            "D13": 4.0,
            "D14": 0.28,
            "D16": 1.00,
            "D9": 2.75,
        }
        
        for name, value in dims.items():
            full = f"{name}@Sketch34"
            dim = swDrawing.Parameter(full)
            if dim:
                dim.SystemValue = value * 0.0254
                log.info(f"  Set Sketch34 dimension {name} to {value:.4f} in ({dim.SystemValue:.6f} m)")
            else:
                log.warning(f"  Sketch34 dimension {name} not found in drawing.")
                
        # Post-write verify
        d4_val = swDrawing.Parameter("D4@Sketch34")
        d5_val = swDrawing.Parameter("D5@Sketch34")
        if d4_val and d5_val:
            d4_in = d4_val.SystemValue / 0.0254
            d5_in = d5_val.SystemValue / 0.0254
            log.info(f"Verified Sketch34 dimensions post-write: D4={d4_in:.2f} (expected {W:.2f}), D5={d5_in:.2f} (expected {H:.2f})")
            assert abs(d4_in - W) < 0.01, f"Sketch34 Width check failed: D4 is {d4_in:.2f} but expected {W:.2f}"
            assert abs(d5_in - H) < 0.01, f"Sketch34 Height check failed: D5 is {d5_in:.2f} but expected {H:.2f}"

    def _reposition_section_line_aa(self, swDrawing):
        """
        Reposition the Section Line (A-A) on Drawing View4 (front view)
        so that it cuts through the rightmost beam/frame edge of the mirror.
        """
        log.info("Starting Section Line A-A repositioning...")
        try:
            swDrawing.ActivateSheet("Sheet1")
            view = swDrawing.GetFirstView
            front_view = None
            while view:
                if view.Name.lower() == "drawing view4":
                    front_view = view
                    break
                view = view.GetNextView
                
            if not front_view:
                log.warning("Drawing View4 (front view) not found on Sheet1. Skipping Section Line A-A repositioning.")
                return

            sec_lines = front_view.GetSectionLines
            if not sec_lines:
                log.warning("No section lines found in Drawing View4. Skipping repositioning.")
                return

            sl = sec_lines[0]
            
            # Retrieve view attributes
            outline = front_view.GetOutline
            position = front_view.Position
            scale_ratio = front_view.ScaleRatio
            
            if not outline or not position or not scale_ratio:
                log.warning("Could not retrieve front view outline, position, or scale. Skipping repositioning.")
                return

            scale = scale_ratio[0] / scale_ratio[1]
            
            # 1. Arrow center at right beam: ~5.8mm inside view right edge on sheet
            arrow_center_x_sheet = outline[2] - 0.0058
            center_x_model = (arrow_center_x_sheet - position[0]) / scale
            
            # 2. Y at vertical midpoint of view
            view_mid_y_sheet = (outline[1] + outline[3]) / 2.0
            center_y_model = (view_mid_y_sheet - position[1]) / scale
            
            # 3. Compact shoulder: ~104mm model half-span = 52.2mm
            half_shoulder = 0.0522 # meters
            
            new_line_info = [
                center_x_model + half_shoulder, center_y_model, 0.0,
                center_x_model - half_shoulder, center_y_model, 0.0
            ]
            
            log.info(f"Setting Section Line A-A to model coords: {new_line_info}")
            
            import win32com.client as win32
            import pythoncom
            dispid = sl._oleobj_.GetIDsOfNames("SetLineInfo")
            var_info = win32.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, new_line_info)
            res = sl._oleobj_.Invoke(dispid, 0, pythoncom.DISPATCH_METHOD, True, var_info)
            
            log.info(f"SetLineInfo returned: {res}")
        except Exception as e:
            log.error(f"Error during Section Line A-A repositioning: {e}", exc_info=True)

    def close(self):
        pass
